import csv
import io
import os
import sqlite3
from datetime import date, datetime
from zipfile import BadZipFile

import click
from flask import (
    Flask,
    Response,
    g,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


app = Flask(__name__)
app.config["DATABASE"] = os.path.join(app.instance_path, "groundtrack.sqlite")


NAV_ITEMS = [
    ("dashboard", "Dashboard"),
    ("import_page", "Import"),
    ("scan", "Scan"),
    ("participants", "Participants"),
    ("walk_up", "Walk-Up"),
    ("on_ground_report", "On-Ground Report"),
]

EXPECTED_IMPORT_COLUMNS = [
    "No.",
    "Name",
    "Rank",
    "NAT",
    "Visit Request Status",
    "Badge #",
    "In-Process Date/Time",
    "Out-Process Date/Time",
    "Your Unit, Organization, or Company",
    "Thread / Initiative",
]

ON_GROUND_EXPORT_HEADERS = [
    "Name",
    "Rank",
    "NAT",
    "Visit Request Status",
    "Badge #",
    "In-Process Date/Time",
    "Unit / Organization / Company",
    "Mission Area / Initiative",
]


@app.context_processor
def inject_navigation():
    return {"nav_items": NAV_ITEMS}


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    os.makedirs(app.instance_path, exist_ok=True)
    db = get_db()

    with app.open_resource("schema.sql") as schema_file:
        db.executescript(schema_file.read().decode("utf-8"))


@app.cli.command("init-db")
def init_db_command():
    init_db()
    click.echo("Initialized the local GroundTrack database.")


def database_exists():
    return os.path.exists(app.config["DATABASE"])


def get_dashboard_counts():
    counts = {
        "total_participants": 0,
        "currently_on_ground": 0,
        "not_checked_in": 0,
        "checked_out": 0,
        "walkup_participants": 0,
    }

    if not database_exists():
        return counts, False

    try:
        row = get_db().execute(
            """
            SELECT
                COUNT(*) AS total_participants,
                SUM(
                    CASE
                        WHEN in_process_at IS NOT NULL
                            AND in_process_at != ''
                            AND (
                                out_process_at IS NULL
                                OR out_process_at = ''
                            )
                        THEN 1 ELSE 0
                    END
                ) AS currently_on_ground,
                SUM(
                    CASE
                        WHEN in_process_at IS NULL
                            OR in_process_at = ''
                        THEN 1 ELSE 0
                    END
                ) AS not_checked_in,
                SUM(
                    CASE
                        WHEN in_process_at IS NOT NULL
                            AND in_process_at != ''
                            AND out_process_at IS NOT NULL
                            AND out_process_at != ''
                        THEN 1 ELSE 0
                    END
                ) AS checked_out,
                SUM(CASE WHEN is_walkup = 1 THEN 1 ELSE 0 END) AS walkup_participants
            FROM participants
            """
        ).fetchone()
    except sqlite3.OperationalError:
        return counts, False

    for key in counts:
        counts[key] = row[key] or 0

    return counts, True


def is_xlsx_file(filename):
    return filename.lower().endswith(".xlsx")


def cell_value_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def get_header_map(sheet):
    header_row = next(
        sheet.iter_rows(min_row=1, max_row=1, values_only=True),
        (),
    )
    return {
        cell_value_text(value): index
        for index, value in enumerate(header_row)
        if cell_value_text(value)
    }


def count_data_rows(sheet, header_map):
    row_count = 0
    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        if any(
            import_cell(row_values, header_map, column)
            for column in EXPECTED_IMPORT_COLUMNS
        ):
            row_count += 1
    return row_count


def import_cell(row_values, header_map, column):
    column_index = header_map[column]
    if column_index >= len(row_values):
        return None

    value = cell_value_text(row_values[column_index])
    return value or None


def open_import_workbook(file_storage):
    try:
        workbook = load_workbook(file_storage, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError):
        return None, None, None, "The uploaded file could not be read as an .xlsx workbook."

    if "Data" not in workbook.sheetnames:
        workbook.close()
        return None, None, None, "The workbook is missing the Data sheet."

    data_sheet = workbook["Data"]
    header_map = get_header_map(data_sheet)
    missing_columns = [
        column for column in EXPECTED_IMPORT_COLUMNS if column not in header_map
    ]

    if missing_columns:
        workbook.close()
        return (
            None,
            None,
            None,
            "Missing expected columns: " + ", ".join(missing_columns),
        )

    return workbook, data_sheet, header_map, None


def validate_import_workbook(file_storage):
    workbook, data_sheet, header_map, error = open_import_workbook(file_storage)
    if error:
        return False, error

    try:
        data_row_count = count_data_rows(data_sheet, header_map)
        return True, f"Spreadsheet validated. Found {data_row_count} data rows."
    finally:
        workbook.close()


def existing_badge_numbers():
    rows = get_db().execute(
        """
        SELECT badge_number
        FROM participants
        WHERE badge_number IS NOT NULL
            AND badge_number != ''
        """
    ).fetchall()
    return {row["badge_number"] for row in rows}


def participant_duplicate_key(participant):
    return (
        participant["name"] or "",
        participant["rank"] or "",
        participant["nat"] or "",
        participant["organization"] or "",
        participant["thread_initiative"] or "",
    )


def existing_participant_duplicate_keys():
    rows = get_db().execute(
        """
        SELECT name, rank, nat, organization, thread_initiative
        FROM participants
        """
    ).fetchall()
    return {
        participant_duplicate_key(
            {
                "name": row["name"],
                "rank": row["rank"],
                "nat": row["nat"],
                "organization": row["organization"],
                "thread_initiative": row["thread_initiative"],
            }
        )
        for row in rows
    }


def row_to_participant(row_values, header_map):
    return {
        "source_row_no": import_cell(row_values, header_map, "No."),
        "name": import_cell(row_values, header_map, "Name"),
        "rank": import_cell(row_values, header_map, "Rank"),
        "nat": import_cell(row_values, header_map, "NAT"),
        "visit_request_status": import_cell(
            row_values,
            header_map,
            "Visit Request Status",
        ),
        "badge_number": import_cell(row_values, header_map, "Badge #"),
        "in_process_at": import_cell(row_values, header_map, "In-Process Date/Time"),
        "out_process_at": import_cell(row_values, header_map, "Out-Process Date/Time"),
        "organization": import_cell(
            row_values,
            header_map,
            "Your Unit, Organization, or Company",
        ),
        "thread_initiative": import_cell(row_values, header_map, "Thread / Initiative"),
        "is_walkup": 0,
    }


def insert_participant(participant):
    db = get_db()
    cursor = db.execute(
        """
        INSERT INTO participants (
            source_row_no,
            name,
            rank,
            nat,
            visit_request_status,
            badge_number,
            in_process_at,
            out_process_at,
            organization,
            thread_initiative,
            is_walkup
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            participant["source_row_no"],
            participant["name"],
            participant["rank"],
            participant["nat"],
            participant["visit_request_status"],
            participant["badge_number"],
            participant["in_process_at"],
            participant["out_process_at"],
            participant["organization"],
            participant["thread_initiative"],
            participant["is_walkup"],
        ),
    )
    db.execute(
        """
        INSERT INTO activity_log (participant_id, action, badge_number, note)
        VALUES (?, ?, ?, ?)
        """,
        (
            cursor.lastrowid,
            "IMPORT",
            participant["badge_number"],
            "Imported from spreadsheet",
        ),
    )


def form_value(name):
    value = request.form.get(name, "").strip()
    return value or None


def blank_walk_up_form():
    return {
        "name": "",
        "rank": "",
        "nat": "",
        "visit_request_status": "",
        "badge_number": "",
        "organization": "",
        "thread_initiative": "",
        "check_in_now": False,
    }


def walk_up_form_data():
    return {
        "name": form_value("name") or "",
        "rank": form_value("rank") or "",
        "nat": form_value("nat") or "",
        "visit_request_status": form_value("visit_request_status") or "",
        "badge_number": form_value("badge_number") or "",
        "organization": form_value("organization") or "",
        "thread_initiative": form_value("thread_initiative") or "",
        "check_in_now": request.form.get("check_in_now") == "1",
    }


def create_walk_up_participant(participant):
    db = get_db()
    in_process_at = local_timestamp() if participant["check_in_now"] else None
    cursor = db.execute(
        """
        INSERT INTO participants (
            name,
            rank,
            nat,
            visit_request_status,
            badge_number,
            organization,
            thread_initiative,
            in_process_at,
            is_walkup
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            participant["name"],
            participant["rank"] or None,
            participant["nat"] or None,
            participant["visit_request_status"] or None,
            participant["badge_number"] or None,
            participant["organization"] or None,
            participant["thread_initiative"] or None,
            in_process_at,
            1,
        ),
    )
    db.execute(
        """
        INSERT INTO activity_log (participant_id, action, badge_number, note)
        VALUES (?, ?, ?, ?)
        """,
        (
            cursor.lastrowid,
            "WALK_UP_CREATE",
            participant["badge_number"] or None,
            "Walk-up participant created",
        ),
    )
    db.commit()


def walk_up_success_redirect(participant):
    query = {"created": participant["name"]}
    if participant["check_in_now"]:
        query["checked_in"] = "1"
    return redirect(url_for("walk_up", **query))


def missing_walk_up_required_fields(form):
    required_fields = [
        ("nat", "NAT"),
        ("badge_number", "Badge #"),
        ("thread_initiative", "Mission Area / Initiative"),
    ]
    return [label for field, label in required_fields if not form[field]]


def import_participants_from_workbook(file_storage):
    workbook, data_sheet, header_map, error = open_import_workbook(file_storage)
    if error:
        return False, error

    imported_count = 0
    skipped_count = 0
    duplicate_skipped_count = 0

    try:
        seen_badges = existing_badge_numbers()
        seen_participant_keys = existing_participant_duplicate_keys()

        for row_values in data_sheet.iter_rows(min_row=2, values_only=True):
            participant = row_to_participant(row_values, header_map)
            badge_number = participant["badge_number"]

            if not participant["name"]:
                skipped_count += 1
                continue

            participant_key = participant_duplicate_key(participant)
            if participant_key in seen_participant_keys:
                skipped_count += 1
                duplicate_skipped_count += 1
                continue

            if badge_number:
                if badge_number in seen_badges:
                    skipped_count += 1
                    duplicate_skipped_count += 1
                    continue
                seen_badges.add(badge_number)

            seen_participant_keys.add(participant_key)
            insert_participant(participant)
            imported_count += 1

        get_db().commit()
    except sqlite3.OperationalError:
        db = g.get("db")
        if db is not None:
            db.rollback()
        return False, "GroundTrack is not ready. Ask the event lead to complete setup."
    finally:
        workbook.close()

    participant_label = "participant" if imported_count == 1 else "participants"
    entry_label = "entry" if skipped_count == 1 else "entries"
    duplicate_label = "duplicate" if duplicate_skipped_count == 1 else "duplicates"
    return (
        True,
        (
            f"Imported {imported_count} {participant_label}. "
            f"Skipped {skipped_count} {entry_label}, including "
            f"{duplicate_skipped_count} {duplicate_label}."
        ),
    )


def local_timestamp():
    return datetime.now().isoformat(sep=" ", timespec="seconds")


def is_currently_on_ground(participant):
    in_process_at = participant["in_process_at"]
    out_process_at = participant["out_process_at"]

    return bool(in_process_at) and (
        not out_process_at or out_process_at < in_process_at
    )


def find_participant_by_badge(badge_number):
    return get_db().execute(
        """
        SELECT id, name, badge_number, in_process_at, out_process_at
        FROM participants
        WHERE badge_number = ?
        ORDER BY id
        LIMIT 1
        """,
        (badge_number,),
    ).fetchone()


def check_in_participant(badge_number):
    try:
        participant = find_participant_by_badge(badge_number)
    except sqlite3.OperationalError:
        return "error", "GroundTrack is not ready. Ask the event lead to complete setup."

    if participant is None:
        return (
            "error",
            f"Badge {badge_number} was not found. Check the badge number and try again.",
        )

    if is_currently_on_ground(participant):
        return (
            "already",
            f"{participant['name']} is already checked in. No changes were made.",
        )

    timestamp = local_timestamp()
    db = get_db()
    db.execute(
        """
        UPDATE participants
        SET in_process_at = ?,
            out_process_at = NULL,
            updated_at = ?
        WHERE id = ?
        """,
        (timestamp, timestamp, participant["id"]),
    )
    db.execute(
        """
        INSERT INTO participant_visits (
            participant_id,
            in_process_at,
            out_process_at
        ) VALUES (?, ?, NULL)
        """,
        (participant["id"], timestamp),
    )
    db.execute(
        """
        INSERT INTO activity_log (participant_id, action, badge_number, note)
        VALUES (?, ?, ?, ?)
        """,
        (
            participant["id"],
            "CHECK_IN",
            participant["badge_number"],
            "Checked in from scan screen",
        ),
    )
    db.commit()

    return "checked_in", f"Checked in {participant['name']}."


def check_out_participant(badge_number):
    try:
        participant = find_participant_by_badge(badge_number)
    except sqlite3.OperationalError:
        return "error", "GroundTrack is not ready. Ask the event lead to complete setup."

    if participant is None:
        return (
            "error",
            f"Badge {badge_number} was not found. Check the badge number and try again.",
        )

    if not is_currently_on_ground(participant):
        return (
            "not_on_ground",
            f"{participant['name']} is not currently checked in. No changes were made.",
        )

    timestamp = local_timestamp()
    db = get_db()
    db.execute(
        """
        UPDATE participants
        SET out_process_at = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (timestamp, timestamp, participant["id"]),
    )
    db.execute(
        """
        INSERT INTO activity_log (participant_id, action, badge_number, note)
        VALUES (?, ?, ?, ?)
        """,
        (
            participant["id"],
            "CHECK_OUT",
            participant["badge_number"],
            "Checked out from scan screen",
        ),
    )
    db.commit()

    return "checked_out", f"Checked out {participant['name']}."


def participant_status(participant):
    if is_currently_on_ground(participant):
        return "On Ground"
    if participant["in_process_at"]:
        return "Off Ground"
    return "Not Checked In"


def get_participants():
    try:
        rows = get_db().execute(
            """
            SELECT
                name,
                rank,
                nat,
                visit_request_status,
                badge_number,
                organization,
                thread_initiative,
                in_process_at,
                out_process_at
            FROM participants
            ORDER BY name COLLATE NOCASE, id
            """
        ).fetchall()
    except sqlite3.OperationalError:
        return []

    participants = []
    for row in rows:
        participant = dict(row)
        participant["status"] = participant_status(row)
        participants.append(participant)

    return participants


def get_on_ground_participants():
    try:
        return get_db().execute(
            """
            SELECT
                name,
                rank,
                nat,
                visit_request_status,
                badge_number,
                in_process_at,
                organization,
                thread_initiative
            FROM participants
            WHERE in_process_at IS NOT NULL
                AND in_process_at != ''
                AND (
                    out_process_at IS NULL
                    OR out_process_at = ''
                )
            ORDER BY name COLLATE NOCASE, id
            """
        ).fetchall()
    except sqlite3.OperationalError:
        return []


def on_ground_export_row(participant):
    return [
        participant["name"],
        participant["rank"] or "",
        participant["nat"] or "",
        participant["visit_request_status"] or "",
        participant["badge_number"] or "",
        participant["in_process_at"],
        participant["organization"] or "",
        participant["thread_initiative"] or "",
    ]


@app.route("/")
def dashboard():
    counts, database_ready = get_dashboard_counts()
    return render_template(
        "dashboard.html",
        counts=counts,
        database_ready=database_ready,
    )


@app.route("/import", methods=["GET", "POST"])
def import_page():
    message = None
    message_type = None

    if request.method == "POST":
        upload = request.files.get("spreadsheet")

        if upload is None or upload.filename == "":
            message = "No file selected. Choose an .xlsx spreadsheet to upload."
            message_type = "error"
        elif not is_xlsx_file(upload.filename):
            message = "File must be an .xlsx spreadsheet."
            message_type = "error"
        else:
            is_imported, message = import_participants_from_workbook(upload)
            message_type = "success" if is_imported else "error"

    return render_template(
        "import.html",
        message=message,
        message_type=message_type,
        expected_columns=EXPECTED_IMPORT_COLUMNS,
    )


@app.route("/scan", methods=["GET", "POST"])
def scan():
    message = None
    message_type = None
    badge_number = ""

    if request.method == "POST":
        badge_number = request.form.get("badge_number", "").strip()
        action = request.form.get("action")

        if not badge_number:
            message = "Enter or scan a badge number."
            message_type = "error"
        elif action == "check_in":
            check_in_status, message = check_in_participant(badge_number)
            message_type = "error" if check_in_status == "error" else "success"
            if check_in_status in {"checked_in", "already"}:
                badge_number = ""
        elif action == "check_out":
            check_out_status, message = check_out_participant(badge_number)
            message_type = "error" if check_out_status == "error" else "success"
            if check_out_status in {"checked_out", "not_on_ground"}:
                badge_number = ""
        else:
            message = "Choose Check In or Check Out."
            message_type = "error"

    return render_template(
        "scan.html",
        message=message,
        message_type=message_type,
        badge_number=badge_number,
    )


@app.route("/participants")
def participants():
    return render_template(
        "participants.html",
        participants=get_participants(),
    )


@app.route("/walk-up", methods=["GET", "POST"])
def walk_up():
    message = None
    message_type = None
    form = blank_walk_up_form()

    if request.args.get("created"):
        if request.args.get("checked_in"):
            message = (
                f"Added {request.args['created']} and checked them in."
            )
        else:
            message = (
                f"Added {request.args['created']}. "
                "They are not checked in."
            )
        message_type = "success"

    if request.method == "POST":
        form = walk_up_form_data()

        if not form["name"]:
            message = "Name is required."
            message_type = "error"
        elif missing_fields := missing_walk_up_required_fields(form):
            missing_fields = ", ".join(missing_fields)
            message = f"Missing required fields: {missing_fields}."
            message_type = "error"
        elif form["badge_number"]:
            try:
                if form["badge_number"] in existing_badge_numbers():
                    message = (
                        f"Badge {form['badge_number']} is already in use. "
                        "Enter a different badge number."
                    )
                    message_type = "error"
                else:
                    create_walk_up_participant(form)
                    return walk_up_success_redirect(form)
            except sqlite3.OperationalError:
                message = (
                    "GroundTrack is not ready. "
                    "Ask the event lead to complete setup."
                )
                message_type = "error"
        else:
            try:
                create_walk_up_participant(form)
                return walk_up_success_redirect(form)
            except sqlite3.OperationalError:
                message = (
                    "GroundTrack is not ready. "
                    "Ask the event lead to complete setup."
                )
                message_type = "error"

    return render_template(
        "walk_up.html",
        form=form,
        message=message,
        message_type=message_type,
    )


@app.route("/on-ground")
def on_ground_report():
    return render_template(
        "on_ground.html",
        participants=get_on_ground_participants(),
    )


@app.route("/on-ground/export.csv")
def export_on_ground_csv():
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(ON_GROUND_EXPORT_HEADERS)

    for participant in get_on_ground_participants():
        writer.writerow(on_ground_export_row(participant))

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=on_ground_report.csv",
        },
    )


@app.route("/on-ground/export.xlsx")
def export_on_ground_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "On-Ground Report"
    sheet.append(ON_GROUND_EXPORT_HEADERS)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for participant in get_on_ground_participants():
        sheet.append(on_ground_export_row(participant))

    column_widths = [28, 12, 12, 22, 16, 24, 32, 28]
    for column_number, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(column_number)].width = width

    sheet.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    return send_file(
        output,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        as_attachment=True,
        download_name="on_ground_report.xlsx",
    )


if __name__ == "__main__":
    app.run(debug=True)
