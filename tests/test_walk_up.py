import csv
import io
import os
import tempfile
import unittest

from openpyxl import load_workbook

from app import app, get_db, init_db


class WalkUpPageTest(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        app.config["DATABASE"] = os.path.join(
            self.tempdir.name,
            "groundtrack.sqlite",
        )

        with app.app_context():
            init_db()

        self.client = app.test_client()

    def tearDown(self):
        self.tempdir.cleanup()

    def add_participant(self, name, badge_number):
        with app.app_context():
            db = get_db()
            db.execute(
                """
                INSERT INTO participants (name, badge_number, is_walkup)
                VALUES (?, ?, ?)
                """,
                (name, badge_number, 0),
            )
            db.commit()

    def test_walk_up_route_shows_form(self):
        response = self.client.get("/walk-up")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Walk-Up", body)
        self.assertIn("Add Walk-Up Participant", body)
        self.assertIn('name="name"', body)
        self.assertIn('name="nat"', body)
        self.assertIn('name="badge_number"', body)
        self.assertIn('name="thread_initiative"', body)
        self.assertIn("Mission Area / Initiative", body)
        self.assertIn('name="check_in_now"', body)
        self.assertIn("Check in now", body)
        self.assertIn("← Back to Dashboard", body)

    def test_valid_walk_up_without_check_in_creates_not_checked_in_participant(self):
        response = self.client.post(
            "/walk-up",
            data={
                "name": " Ada Lovelace ",
                "rank": " CPT ",
                "nat": " US ",
                "visit_request_status": "Approved",
                "badge_number": " W100 ",
                "organization": " Example Unit ",
                "thread_initiative": " Alpha ",
            },
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Added Ada Lovelace. They are not checked in.", body)

        with app.app_context():
            row = get_db().execute(
                """
                SELECT name, rank, nat, visit_request_status, badge_number,
                    organization, thread_initiative, in_process_at,
                    out_process_at, is_walkup
                FROM participants
                WHERE badge_number = ?
                """,
                ("W100",),
            ).fetchone()

        self.assertIsNotNone(row)
        self.assertEqual(row["name"], "Ada Lovelace")
        self.assertEqual(row["rank"], "CPT")
        self.assertEqual(row["nat"], "US")
        self.assertEqual(row["visit_request_status"], "Approved")
        self.assertEqual(row["organization"], "Example Unit")
        self.assertEqual(row["thread_initiative"], "Alpha")
        self.assertIsNone(row["in_process_at"])
        self.assertIsNone(row["out_process_at"])
        self.assertEqual(row["is_walkup"], 1)

        with app.app_context():
            open_visit_count = get_db().execute(
                """
                SELECT COUNT(*) AS count
                FROM participant_visits
                WHERE participant_id = (
                    SELECT id FROM participants WHERE badge_number = ?
                )
                    AND out_process_at IS NULL
                """,
                ("W100",),
            ).fetchone()["count"]

        self.assertEqual(open_visit_count, 0)

    def test_valid_walk_up_with_check_in_sets_in_process_at_and_open_visit(self):
        response = self.client.post(
            "/walk-up",
            data={
                "name": "Katherine Johnson",
                "nat": "US",
                "badge_number": "W101",
                "thread_initiative": "Alpha",
                "check_in_now": "1",
            },
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn(
            "Added Katherine Johnson and checked them in.",
            body,
        )

        with app.app_context():
            row = get_db().execute(
                """
                SELECT id, in_process_at, out_process_at, is_walkup
                FROM participants
                WHERE badge_number = ?
                """,
                ("W101",),
            ).fetchone()
            visits = get_db().execute(
                """
                SELECT participant_id, in_process_at, out_process_at
                FROM participant_visits
                WHERE participant_id = ?
                ORDER BY id
                """,
                (row["id"],),
            ).fetchall()

        self.assertIsNotNone(row)
        self.assertIsNotNone(row["in_process_at"])
        self.assertIsNone(row["out_process_at"])
        self.assertEqual(row["is_walkup"], 1)
        self.assertEqual(len(visits), 1)
        self.assertEqual(visits[0]["participant_id"], row["id"])
        self.assertEqual(visits[0]["in_process_at"], row["in_process_at"])
        self.assertIsNone(visits[0]["out_process_at"])

    def test_checked_in_walk_up_detail_shows_on_ground_and_open_visit(self):
        self.client.post(
            "/walk-up",
            data={
                "name": "Annie Easley",
                "nat": "US",
                "badge_number": "W104",
                "thread_initiative": "Delta",
                "check_in_now": "1",
            },
        )

        with app.app_context():
            row = get_db().execute(
                """
                SELECT id, in_process_at
                FROM participants
                WHERE badge_number = ?
                """,
                ("W104",),
            ).fetchone()

        response = self.client.get(f"/participants/{row['id']}")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Current status: On Ground", body)
        self.assertIn(row["in_process_at"], body)
        self.assertIn("<td>Open</td>", body)

    def test_checked_in_walk_up_appears_on_ground_report_and_exports(self):
        self.client.post(
            "/walk-up",
            data={
                "name": "Dorothy Vaughan",
                "rank": "CIV",
                "nat": "US",
                "visit_request_status": "Approved",
                "badge_number": "W105",
                "organization": "Example Unit",
                "thread_initiative": "Echo",
                "check_in_now": "1",
            },
        )

        report_response = self.client.get("/on-ground")
        csv_response = self.client.get("/on-ground/export.csv")
        excel_response = self.client.get("/on-ground/export.xlsx")

        self.assertEqual(report_response.status_code, 200)
        self.assertIn("Dorothy Vaughan", report_response.get_data(as_text=True))

        self.assertEqual(csv_response.status_code, 200)
        csv_rows = list(csv.reader(io.StringIO(csv_response.get_data(as_text=True))))
        self.assertEqual(csv_rows[1][0], "Dorothy Vaughan")
        self.assertEqual(csv_rows[1][4], "W105")

        self.assertEqual(excel_response.status_code, 200)
        workbook = load_workbook(io.BytesIO(excel_response.data), read_only=True)
        sheet = workbook["On-Ground Report"]
        excel_rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        self.assertEqual(excel_rows[1][0], "Dorothy Vaughan")
        self.assertEqual(excel_rows[1][4], "W105")

    def test_created_walk_up_appears_on_participants_page(self):
        self.client.post(
            "/walk-up",
            data={
                "name": "Grace Hopper",
                "nat": "US",
                "badge_number": "W102",
                "thread_initiative": "Bravo",
            },
        )

        response = self.client.get("/participants")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Grace Hopper", body)
        self.assertIn("Not Checked In", body)

    def test_checked_in_walk_up_appears_on_ground_on_participants_page(self):
        self.client.post(
            "/walk-up",
            data={
                "name": "Mary Jackson",
                "nat": "US",
                "badge_number": "W103",
                "thread_initiative": "Charlie",
                "check_in_now": "1",
            },
        )

        response = self.client.get("/participants")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Mary Jackson", body)
        self.assertIn("On Ground", body)

    def test_blank_name_is_rejected(self):
        response = self.client.post(
            "/walk-up",
            data={
                "name": "   ",
                "nat": "US",
                "badge_number": "W200",
                "thread_initiative": "Delta",
            },
        )

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Name is required.", body)

        with app.app_context():
            count = get_db().execute(
                "SELECT COUNT(*) AS count FROM participants"
            ).fetchone()["count"]

        self.assertEqual(count, 0)

    def test_duplicate_nonblank_badge_is_rejected(self):
        self.add_participant("Existing Person", "W300")

        response = self.client.post(
            "/walk-up",
            data={
                "name": "New Person",
                "nat": "US",
                "badge_number": "W300",
                "thread_initiative": "Echo",
            },
        )

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn(
            "Badge W300 is already in use. Enter a different badge number.",
            body,
        )

        with app.app_context():
            count = get_db().execute(
                "SELECT COUNT(*) AS count FROM participants"
            ).fetchone()["count"]

        self.assertEqual(count, 1)

    def test_required_nat_badge_and_mission_area_are_rejected(self):
        response = self.client.post(
            "/walk-up",
            data={
                "name": "New Person",
                "nat": " ",
                "badge_number": "",
                "thread_initiative": " ",
            },
        )

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn(
            "Missing required fields: NAT, Badge #, Mission Area / Initiative.",
            body,
        )

        with app.app_context():
            count = get_db().execute(
                "SELECT COUNT(*) AS count FROM participants"
            ).fetchone()["count"]

        self.assertEqual(count, 0)


if __name__ == "__main__":
    unittest.main()
