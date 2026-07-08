import csv
import io
import os
import tempfile
import unittest

from openpyxl import load_workbook

from app import app, get_db, init_db


class OnGroundReportTest(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        app.config["DATABASE"] = os.path.join(
            self.tempdir.name,
            "groundtrack.sqlite",
        )

        with app.app_context():
            init_db()

        self.client = app.test_client()

    def tearDown(self):
        self.tempdir.cleanup()

    def add_participant(
        self,
        name,
        badge_number,
        in_process_at=None,
        out_process_at=None,
    ):
        with app.app_context():
            db = get_db()
            db.execute(
                """
                INSERT INTO participants (
                    name,
                    rank,
                    nat,
                    visit_request_status,
                    badge_number,
                    organization,
                    thread_initiative,
                    in_process_at,
                    out_process_at,
                    is_walkup
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    name,
                    "CPT",
                    "US",
                    "Approved",
                    badge_number,
                    "Example Unit",
                    "Alpha",
                    in_process_at,
                    out_process_at,
                    0,
                ),
            )
            participant_id = db.execute(
                "SELECT last_insert_rowid()"
            ).fetchone()[0]
            db.commit()
            return participant_id

    def add_visit(self, participant_id, in_process_at, out_process_at=None):
        with app.app_context():
            db = get_db()
            db.execute(
                """
                INSERT INTO participant_visits (
                    participant_id,
                    in_process_at,
                    out_process_at
                ) VALUES (?, ?, ?)
                """,
                (participant_id, in_process_at, out_process_at),
            )
            db.commit()

    def add_badge_history(self, participant_id, old_badge, new_badge):
        with app.app_context():
            db = get_db()
            db.execute(
                """
                INSERT INTO badge_history (
                    participant_id,
                    old_badge,
                    new_badge,
                    changed_at
                ) VALUES (?, ?, ?, ?)
                """,
                (
                    participant_id,
                    old_badge,
                    new_badge,
                    "2026-07-07 12:00:00",
                ),
            )
            db.commit()

    def test_on_ground_route_empty_state(self):
        response = self.client.get("/on-ground")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("On-Ground Report", body)
        self.assertIn("Nobody is currently on ground.", body)
        self.assertIn("← Back to Dashboard", body)
        self.assertIn('href="/on-ground/export.csv"', body)
        self.assertIn("Export CSV", body)
        self.assertIn('href="/on-ground/export.xlsx"', body)
        self.assertIn("Export Excel", body)
        self.assertNotIn("Thread / Initiative", body)

    def test_on_ground_route_only_shows_currently_on_ground(self):
        self.add_participant(
            "On Ground Person",
            "1001",
            in_process_at="2026-07-07 09:00:00",
            out_process_at=None,
        )
        self.add_participant(
            "Checked Out Person",
            "1002",
            in_process_at="2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )
        self.add_participant("Not Checked In Person", "1003")

        response = self.client.get("/on-ground")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("On Ground Person", body)
        self.assertIn("1001", body)
        self.assertIn("2026-07-07 09:00:00", body)
        self.assertIn("Example Unit", body)
        self.assertIn("Alpha", body)
        self.assertIn("Mission Area / Initiative", body)
        self.assertNotIn("Checked Out Person", body)
        self.assertNotIn("Not Checked In Person", body)

    def test_on_ground_route_excludes_blank_out_process_only_when_checked_in(self):
        self.add_participant(
            "Blank Out Process Person",
            "1004",
            in_process_at="2026-07-07 11:00:00",
            out_process_at="",
        )

        response = self.client.get("/on-ground")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Blank Out Process Person", body)

    def test_on_ground_route_uses_open_visit_status_and_timestamp(self):
        participant_id = self.add_participant(
            "Return Visit Person",
            "2001",
            in_process_at="2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )
        self.add_visit(
            participant_id,
            "2026-07-08 08:30:00",
            out_process_at=None,
        )

        response = self.client.get("/on-ground")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Return Visit Person", body)
        self.assertIn("2026-07-08 08:30:00", body)
        self.assertNotIn("2026-07-07 09:00:00", body)

    def test_on_ground_route_excludes_closed_visit_when_not_on_ground(self):
        participant_id = self.add_participant(
            "Closed Visit Person",
            "2002",
            in_process_at="2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )
        self.add_visit(
            participant_id,
            "2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )

        response = self.client.get("/on-ground")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertNotIn("Closed Visit Person", body)

    def test_on_ground_route_includes_legacy_on_ground_without_visit(self):
        self.add_participant(
            "Legacy On Ground Person",
            "2003",
            in_process_at="2026-07-07 09:00:00",
            out_process_at=None,
        )

        response = self.client.get("/on-ground")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Legacy On Ground Person", body)
        self.assertIn("2026-07-07 09:00:00", body)

    def test_on_ground_route_shows_current_badge_after_replacement(self):
        participant_id = self.add_participant(
            "Replacement Badge Person",
            "3002",
            in_process_at="2026-07-07 09:00:00",
            out_process_at=None,
        )
        self.add_visit(
            participant_id,
            "2026-07-07 09:00:00",
            out_process_at=None,
        )
        self.add_badge_history(participant_id, "3001", "3002")

        response = self.client.get("/on-ground")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("Replacement Badge Person", body)
        self.assertIn("3002", body)
        self.assertNotIn("3001", body)

    def test_on_ground_route_does_not_duplicate_participant_for_visits_or_badge_history(self):
        participant_id = self.add_participant(
            "Duplicate Guard Person",
            "4003",
            in_process_at="2026-07-07 09:00:00",
            out_process_at=None,
        )
        self.add_visit(
            participant_id,
            "2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )
        self.add_visit(
            participant_id,
            "2026-07-08 09:00:00",
            out_process_at=None,
        )
        self.add_visit(
            participant_id,
            "2026-07-09 09:00:00",
            out_process_at=None,
        )
        self.add_badge_history(participant_id, "4001", "4002")
        self.add_badge_history(participant_id, "4002", "4003")

        response = self.client.get("/on-ground")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertEqual(body.count("Duplicate Guard Person"), 1)
        self.assertIn("2026-07-09 09:00:00", body)
        self.assertNotIn("2026-07-08 09:00:00", body)

    def test_export_returns_only_currently_on_ground_participants(self):
        self.add_participant(
            "On Ground Person",
            "1001",
            in_process_at="2026-07-07 09:00:00",
            out_process_at=None,
        )
        self.add_participant(
            "Checked Out Person",
            "1002",
            in_process_at="2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )
        self.add_participant("Not Checked In Person", "1003")

        response = self.client.get("/on-ground/export.csv")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "text/csv")
        self.assertIn(
            "attachment; filename=on_ground_report.csv",
            response.headers["Content-Disposition"],
        )
        rows = list(csv.reader(io.StringIO(response.get_data(as_text=True))))
        self.assertEqual(
            rows[0],
            [
                "Name",
                "Rank",
                "NAT",
                "Visit Request Status",
                "Badge #",
                "In-Process Date/Time",
                "Unit / Organization / Company",
                "Mission Area / Initiative",
            ],
        )
        self.assertEqual(
            rows[1],
            [
                "On Ground Person",
                "CPT",
                "US",
                "Approved",
                "1001",
                "2026-07-07 09:00:00",
                "Example Unit",
                "Alpha",
            ],
        )
        self.assertEqual(len(rows), 2)

    def test_csv_export_uses_open_visit_timestamp(self):
        participant_id = self.add_participant(
            "Return Visit Person",
            "2001",
            in_process_at="2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )
        self.add_visit(
            participant_id,
            "2026-07-08 08:30:00",
            out_process_at=None,
        )

        response = self.client.get("/on-ground/export.csv")

        self.assertEqual(response.status_code, 200)
        rows = list(csv.reader(io.StringIO(response.get_data(as_text=True))))
        self.assertEqual(rows[1][0], "Return Visit Person")
        self.assertEqual(rows[1][5], "2026-07-08 08:30:00")

    def test_empty_export_returns_headers_only(self):
        response = self.client.get("/on-ground/export.csv")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "text/csv")
        rows = list(csv.reader(io.StringIO(response.get_data(as_text=True))))
        self.assertEqual(len(rows), 1)
        self.assertEqual(
            rows[0],
            [
                "Name",
                "Rank",
                "NAT",
                "Visit Request Status",
                "Badge #",
                "In-Process Date/Time",
                "Unit / Organization / Company",
                "Mission Area / Initiative",
            ],
        )

    def test_excel_export_returns_only_currently_on_ground_participants(self):
        self.add_participant(
            "On Ground Person",
            "1001",
            in_process_at="2026-07-07 09:00:00",
            out_process_at=None,
        )
        self.add_participant(
            "Checked Out Person",
            "1002",
            in_process_at="2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )
        self.add_participant("Not Checked In Person", "1003")

        response = self.client.get("/on-ground/export.xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "attachment; filename=on_ground_report.xlsx",
            response.headers["Content-Disposition"],
        )

        workbook = load_workbook(io.BytesIO(response.data))
        sheet = workbook["On-Ground Report"]
        rows = list(sheet.iter_rows(values_only=True))
        self.assertTrue(sheet["A1"].font.bold)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertGreater(sheet.column_dimensions["A"].width, 10)
        workbook.close()

        self.assertEqual(
            rows[0],
            (
                "Name",
                "Rank",
                "NAT",
                "Visit Request Status",
                "Badge #",
                "In-Process Date/Time",
                "Unit / Organization / Company",
                "Mission Area / Initiative",
            ),
        )
        self.assertEqual(
            rows[1],
            (
                "On Ground Person",
                "CPT",
                "US",
                "Approved",
                "1001",
                "2026-07-07 09:00:00",
                "Example Unit",
                "Alpha",
            ),
        )
        self.assertEqual(len(rows), 2)

    def test_excel_export_uses_open_visit_timestamp(self):
        participant_id = self.add_participant(
            "Return Visit Person",
            "2001",
            in_process_at="2026-07-07 09:00:00",
            out_process_at="2026-07-07 10:00:00",
        )
        self.add_visit(
            participant_id,
            "2026-07-08 08:30:00",
            out_process_at=None,
        )

        response = self.client.get("/on-ground/export.xlsx")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        sheet = workbook["On-Ground Report"]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        self.assertEqual(rows[1][0], "Return Visit Person")
        self.assertEqual(rows[1][5], "2026-07-08 08:30:00")

    def test_empty_excel_export_returns_headers_only(self):
        response = self.client.get("/on-ground/export.xlsx")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        sheet = workbook["On-Ground Report"]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        self.assertEqual(len(rows), 1)
        self.assertEqual(
            rows[0],
            (
                "Name",
                "Rank",
                "NAT",
                "Visit Request Status",
                "Badge #",
                "In-Process Date/Time",
                "Unit / Organization / Company",
                "Mission Area / Initiative",
            ),
        )


if __name__ == "__main__":
    unittest.main()
